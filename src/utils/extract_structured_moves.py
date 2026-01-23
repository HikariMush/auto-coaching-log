#!/usr/bin/env python3
"""
スマブラSP フレームデータの構造化抽出
技データを個別のフィールドに分解してSQLiteに保存
"""
import os
import re
import sqlite3
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd
import openpyxl

EXCEL_FILE = Path(__file__).parent.parent / 'brain' / 'raw_data' / 'スマブラSP フレームデータ by検証窓.xlsx'
DB_FILE = Path(__file__).parent.parent.parent / 'data' / 'framedata.db'


def create_database():
    """技データ用のSQLiteデータベースを作成"""
    
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # charactersテーブル
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS characters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            sheet_name TEXT
        )
    ''')
    
    # movesテーブル（技データ）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS moves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            char_id INTEGER NOT NULL,
            move_name TEXT NOT NULL,
            move_category TEXT,
            startup INTEGER,
            active_frames TEXT,
            total_frames INTEGER,
            base_damage REAL,
            damage_1v1 REAL,
            landing_lag INTEGER,
            shield_hitstun INTEGER,
            shield_advantage INTEGER,
            note TEXT,
            raw_data TEXT,
            FOREIGN KEY (char_id) REFERENCES characters (id),
            UNIQUE(char_id, move_name, move_category)
        )
    ''')
    
    # インデックス
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_char_id ON moves(char_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_move_name ON moves(move_name)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_char_move ON moves(char_id, move_name)')
    
    conn.commit()
    conn.close()
    
    print(f"✅ データベース作成完了: {DB_FILE}")


def extract_number(value: str) -> Optional[int]:
    """文字列から最初の整数を抽出"""
    if pd.isna(value):
        return None
    
    match = re.search(r'(\d+)', str(value))
    if match:
        return int(match.group(1))
    return None


def extract_float(value: str) -> Optional[float]:
    """文字列から最初の浮動小数点数を抽出"""
    if pd.isna(value):
        return None
    
    match = re.search(r'(\d+\.?\d*)', str(value))
    if match:
        return float(match.group(1))
    return None


def parse_ground_moves(df: pd.DataFrame, start_row: int) -> List[Dict]:
    """
    地上攻撃セクションをパース
    
    返り値の例:
    [
        {
            'move_name': '弱1',
            'move_category': '弱攻撃',
            'startup': 2,
            'active_frames': '2-4',
            'total_frames': 20,
            'base_damage': 2.0,
            'damage_1v1': 2.4,
            'shield_advantage': '3F',
            'note': '...'
        },
        ...
    ]
    """
    moves = []
    
    # ヘッダー行を探す
    header_row = None
    for i in range(start_row, min(start_row + 5, len(df))):
        row = df.iloc[i]
        if '判定持続' in str(row.values):
            header_row = i
            break
    
    if header_row is None:
        return moves
    
    # データ行を処理
    current_category = None
    for i in range(header_row + 1, len(df)):
        row = df.iloc[i]
        
        # 次のセクションに到達したら終了
        if pd.notna(row[0]) and any(keyword in str(row[0]) for keyword in ['空中攻撃', '必殺ワザ', 'つかみ']):
            break
        
        # カテゴリ行（例: '弱攻撃', '強攻撃'）
        if pd.notna(row[0]) and pd.isna(row[2]) and len(str(row[0])) < 10:
            current_category = str(row[0])
            continue
        
        # データ行
        if pd.notna(row[1]):  # サブ技名がある（弱1, 弱2など）
            move_name = str(row[1])
        elif pd.notna(row[0]):  # メイン技名のみ
            move_name = str(row[0])
        else:
            continue
        
        # 発生フレームを判定持続から抽出
        active_str = str(row[2]) if pd.notna(row[2]) else None
        startup = None
        if active_str:
            startup = extract_number(active_str.split('-')[0])
        
        # 数値データを抽出
        move_data = {
            'move_name': move_name,
            'move_category': current_category or '地上攻撃',
            'startup': startup,
            'active_frames': active_str,
            'total_frames': extract_number(row[3]),
            'base_damage': extract_float(row[4]),
            'damage_1v1': extract_float(row[5]),
            'shield_hitstun': extract_number(row[6]),  # ガード硬直
            'note': str(row[7]) if pd.notna(row[7]) else None,
        }
        
        # ガード硬直差を計算：全体F - 発生F - ガード硬直F
        if move_data['total_frames'] and move_data['startup'] and move_data['shield_hitstun']:
            shield_adv = move_data['total_frames'] - move_data['startup'] - move_data['shield_hitstun']
            move_data['shield_advantage'] = shield_adv
        
        moves.append(move_data)
    
    return moves


def parse_aerial_moves(df: pd.DataFrame, start_row: int) -> List[Dict]:
    """空中攻撃セクションをパース"""
    moves = []
    
    # ヘッダー行を探す
    header_row = None
    for i in range(start_row, min(start_row + 5, len(df))):
        row = df.iloc[i]
        if '判定持続' in str(row.values):
            header_row = i
            break
    
    if header_row is None:
        return moves
    
    # データ行を処理
    for i in range(header_row + 1, len(df)):
        row = df.iloc[i]
        
        # 次のセクションに到達したら終了
        if pd.notna(row[0]) and any(keyword in str(row[0]) for keyword in ['必殺ワザ', 'つかみ', '回避']):
            break
        
        # 技名
        if pd.notna(row[0]):
            move_name = str(row[0])
            if move_name in ['NaN', 'nan'] or len(move_name) > 20:
                continue
        else:
            continue
        
        # サブ技（持続、連など）
        sub_name = str(row[1]) if pd.notna(row[1]) else None
        if sub_name and sub_name not in ['NaN', 'nan']:
            move_name = f"{move_name}_{sub_name}"
        
        # 発生フレームを判定持続から抽出
        active_str = str(row[2]) if pd.notna(row[2]) else None
        startup = None
        if active_str:
            startup = extract_number(active_str.split('-')[0])
        
        # 数値データを抽出
        move_data = {
            'move_name': move_name,
            'move_category': '空中攻撃',
            'startup': startup,
            'active_frames': active_str,
            'total_frames': extract_number(row[3]),
            'base_damage': extract_float(row[4]),
            'damage_1v1': extract_float(row[5]),
            'shield_hitstun': extract_number(row[7]),  # ガード硬直（列7）
            'landing_lag': extract_number(row[8]),
            'note': str(row[17]) if pd.notna(row[17]) else None,
        }
        
        # 空中技のガード硬直差を計算
        # 空中技の場合: 着地隙 - ガード硬直F （負の値なら攻撃側が不利）
        if move_data['landing_lag'] is not None and move_data['shield_hitstun'] is not None:
            shield_adv = move_data['landing_lag'] - move_data['shield_hitstun']
            move_data['shield_advantage'] = shield_adv
        
        moves.append(move_data)
    
    return moves


def extract_character_moves(sheet_name: str, wb: openpyxl.Workbook) -> Tuple[str, List[Dict]]:
    """
    1キャラクターのシートから技データを抽出
    
    Returns:
        (character_name, moves_list)
    """
    # キャラクター名を抽出
    match = re.search(r'[0-9０-９]+\.\s*(.+)', sheet_name)
    character_name = match.group(1) if match else sheet_name
    
    # シートをDataFrameに変換
    sheet = wb[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    df = pd.DataFrame(data)
    
    all_moves = []
    
    # 地上攻撃セクションを探す
    for i in range(len(df)):
        if '地上攻撃' in str(df.iloc[i, 0]):
            ground_moves = parse_ground_moves(df, i)
            all_moves.extend(ground_moves)
            break
    
    # 空中攻撃セクションを探す
    for i in range(len(df)):
        if '空中攻撃' in str(df.iloc[i, 0]):
            aerial_moves = parse_aerial_moves(df, i)
            all_moves.extend(aerial_moves)
            break
    
    return character_name, all_moves


def save_to_database(character_name: str, moves: List[Dict]):
    """キャラクターと技データをデータベースに保存"""
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # キャラクター登録
    cursor.execute('''
        INSERT OR IGNORE INTO characters (name, sheet_name)
        VALUES (?, ?)
    ''', (character_name, character_name))
    
    # キャラクターIDを取得
    cursor.execute('SELECT id FROM characters WHERE name = ?', (character_name,))
    char_id = cursor.fetchone()[0]
    
    # 技データを登録
    for move in moves:
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO moves
                (char_id, move_name, move_category, startup, active_frames,
                 total_frames, base_damage, damage_1v1, landing_lag,
                 shield_hitstun, shield_advantage, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                char_id,
                move.get('move_name'),
                move.get('move_category'),
                move.get('startup'),
                move.get('active_frames'),
                move.get('total_frames'),
                move.get('base_damage'),
                move.get('damage_1v1'),
                move.get('landing_lag'),
                move.get('shield_hitstun'),
                move.get('shield_advantage'),
                move.get('note')
            ))
        except Exception as e:
            print(f"  ⚠️ 技登録エラー: {move.get('move_name')} - {e}")
    
    conn.commit()
    conn.close()


def main():
    """メイン処理"""
    
    print("="*70)
    print("📊 スマブラSP フレームデータ構造化抽出")
    print("="*70 + "\n")
    
    # データベース作成
    create_database()
    
    # Excelファイル読み込み
    if not EXCEL_FILE.exists():
        print(f"❌ Excelファイルが見つかりません: {EXCEL_FILE}")
        return
    
    print(f"📂 Excelファイル読み込み中: {EXCEL_FILE.name}\n")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # 各キャラクターシートを処理
    total_moves = 0
    for sheet_idx, sheet_name in enumerate(wb.sheetnames[1:], start=1):
        # キャラクターシートのみ処理（番号付きシート）
        if not re.match(r'^[0-9０-９]+\.', sheet_name):
            continue
        
        print(f"[{sheet_idx:3d}] {sheet_name}")
        
        try:
            character_name, moves = extract_character_moves(sheet_name, wb)
            
            if moves:
                save_to_database(character_name, moves)
                print(f"     ✅ {len(moves)}技を登録")
                total_moves += len(moves)
            else:
                print(f"     ⚠️ 技データが見つかりませんでした")
        
        except Exception as e:
            print(f"     ❌ エラー: {e}")
    
    print("\n" + "="*70)
    print(f"✅ 完了: 合計 {total_moves} 技を登録")
    print(f"📋 データベース: {DB_FILE}")
    print("="*70)


if __name__ == '__main__':
    main()
