#!/usr/bin/env python3
"""
SmashZettel-Bot: Technique Data Ingestion from Excel + Text Data
Combines Excel frame data with raw_data text files for comprehensive technique knowledge base.

Usage:
    python -m src.utils.ingest_excel_data [--resume] [--dry-run]
"""

import os
import json
import sys
import time
import argparse
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import re

import openpyxl
import google.generativeai as genai
from pinecone import Pinecone

# Configuration
EXCEL_FILE = Path(__file__).parent.parent / 'brain' / 'raw_data' / 'スマブラSP フレームデータ by検証窓.xlsx'
RAW_DATA_DIR = Path(__file__).parent.parent / 'brain' / 'raw_data'
DATA_DIR = Path(__file__).parent.parent.parent / 'data'
INGESTION_STATE_FILE = DATA_DIR / 'excel_ingestion_state.json'


def initialize_apis() -> Tuple[Any, Pinecone]:
    """Initialize Google Generative AI and Pinecone"""
    
    # Initialize Gemini
    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        raise ValueError("GEMINI_API_KEY environment variable not set")
    genai.configure(api_key=api_key)
    
    # Initialize Pinecone
    pinecone_key = os.environ.get('PINECONE_API_KEY')
    if not pinecone_key:
        raise ValueError("PINECONE_API_KEY environment variable not set")
    
    pc = Pinecone(api_key=pinecone_key)
    return genai, pc


def load_excel_workbook() -> openpyxl.Workbook:
    """Load Excel workbook"""
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    
    print(f"📂 Loading Excel file: {EXCEL_FILE.name}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)  # 数式の結果値を取得
    print(f"✅ Found {len(wb.sheetnames)} character sheets\n")
    
    return wb


def extract_excel_sections(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                          character_name: str) -> Dict[str, List[str]]:
    """
    Extract structured sections from Excel sheet
    
    Returns:
        {
            '行動フレーム': ['Jab data...', 'Up-tilt data...', ...],
            '能力値': ['Weight: 98', 'Fall speed: 1.6', ...],
            '滞空フレーム': ['Jump: 45', 'Air time: 12', ...],
        }
    """
    
    sections = {
        '行動フレーム': [],
        '能力値': [],
        '滞空フレーム': [],
    }
    
    # Map headers to section names
    section_keywords = {
        '行動フレーム': '行動フレーム',
        '能力値': '能力値',
        '滞空フレーム': '滞空フレーム',
    }
    
    current_section = None
    current_buffer = []
    
    # Scan all rows
    for row_idx in range(1, sheet.max_row + 1):
        row_data = []
        
        for col_idx in range(1, min(sheet.max_column + 1, 15)):  # First 15 columns
            cell = sheet.cell(row_idx, col_idx)
            if cell.value is None:
                continue
            
            value = str(cell.value)
            
            # Skip formulas and images
            if value.startswith("='") or value.startswith('=') or '<' in value:
                continue
            
            row_data.append(value.strip())
        
        if not row_data:
            continue
        
        row_text = ' | '.join(row_data)
        
        # Check for section header
        for section_key, header in section_keywords.items():
            if header in row_text:
                # Save previous buffer
                if current_buffer and current_section:
                    sections[current_section].append(' | '.join(current_buffer))
                
                current_section = section_key
                current_buffer = []
                break
        else:
            # Add to current buffer
            if current_section and len(row_data) >= 2:
                current_buffer.extend(row_data)
                
                # If buffer is getting long, flush it
                if len(current_buffer) > 20:
                    sections[current_section].append(' | '.join(current_buffer))
                    current_buffer = []
    
    # Flush remaining buffer
    if current_buffer and current_section:
        sections[current_section].append(' | '.join(current_buffer))
    
    # Clean up sections
    for section in sections:
        # Remove duplicates and empty strings
        sections[section] = [s.strip() for s in sections[section] if s.strip()]
    
    return sections


def load_raw_text_data() -> Dict[str, str]:
    """Load raw text data from .txt files"""
    
    text_data = {}
    
    if not RAW_DATA_DIR.exists():
        return text_data
    
    for txt_file in RAW_DATA_DIR.glob('*.txt'):
        section_name = txt_file.stem  # Filename without .txt
        try:
            with open(txt_file, 'r', encoding='utf-8') as f:
                content = f.read()
                text_data[section_name] = content
        except Exception as e:
            print(f"⚠️  Could not read {txt_file}: {e}")
    
    return text_data


def format_technique_text(character: str, section: str, data: str) -> str:
    """Format technique data as structured text for embedding"""
    
    lines = [
        f"【キャラクター】{character}",
        f"【カテゴリ】{section}",
        f"【データ】{data[:500]}",  # Limit to 500 chars
    ]
    
    return "\n".join(lines)


def embed_text(genai_client: Any, text: str, delay: float = 0.5) -> Optional[List[float]]:
    """Generate embedding for text using Gemini embedding-001"""
    
    try:
        time.sleep(delay)  # Rate limit protection
        response = genai_client.embed_content(
            model="models/embedding-001",
            content=text,
            task_type="SEMANTIC_SIMILARITY"
        )
        return response['embedding']
    except Exception as e:
        print(f"⚠️  Embedding failed: {e}")
        return None


def generate_section_metadata(genai_client: Any, character: str, section_name: str,
                              entries_preview: str, delay: float = 1.0) -> Dict[str, Any]:
    """
    Quick LLM analysis for a single section's techniques
    Returns structured metadata for DSPy use
    """
    
    try:
        time.sleep(delay)  # Rate limit protection
        prompt = f"""
キャラクター: {character}
セクション: {section_name}

技データ:
{entries_preview}

JSON形式で分析結果を返してください:
{{
  "section_type": "弱攻撃/強攻撃など",
  "common_damage_range": "ダメージ幅",
  "avg_startup": "平均発生",
  "general_use": "一般的な用途",
  "combo_rating": "high/medium/low"
}}

JSON形式のみ。
"""
        
        model = genai.GenerativeModel("gemini-2.5-flash")
        response = model.generate_content(prompt, generation_config={"max_output_tokens": 300})
        
        json_str = response.text.strip()
        if json_str.startswith("```"):
            json_str = json_str.split("```")[1]
            if json_str.startswith("json"):
                json_str = json_str[4:]
            json_str = json_str.strip()
        
        return json.loads(json_str)
    
    except Exception as e:
        return {
            'section_type': section_name,
            'common_damage_range': '不明',
            'avg_startup': '不明',
            'general_use': '技データ',
            'combo_rating': 'medium'
        }


def save_to_pinecone(pc: Pinecone, character: str, section: str, data: str,
                     embedding: List[float], batch_id: int,
                     llm_metadata: Optional[Dict[str, Any]] = None) -> bool:
    """Save technique data to Pinecone with LLM metadata"""
    
    try:
        index_name = "smash-coach-index"
        try:
            index = pc.Index(index_name)
        except:
            return False
        
        # Create unique ID (ASCII only for Pinecone)
        import hashlib
        char_hash = hashlib.md5(character.encode('utf-8')).hexdigest()[:8]
        section_hash = hashlib.md5(section.encode('utf-8')).hexdigest()[:8]
        vector_id = f"excel_{char_hash}_{section_hash}_{batch_id}"
        
        # Base metadata
        metadata = {
            'character': character,
            'section': section,
            'source': 'excel_ingestion',
            'data_preview': data[:300],
        }
        
        # Add LLM metadata if available
        if llm_metadata:
            metadata.update({
                'attack_type': str(llm_metadata.get('attack_type', section))[:100],
                'damage': str(llm_metadata.get('damage', '不明'))[:50],
                'startup': str(llm_metadata.get('startup', '不明'))[:30],
                'combo_potential': str(llm_metadata.get('combo_potential', 'unknown'))[:20],
                'role': str(llm_metadata.get('role', '不明'))[:100],
                'dspy_context': str(llm_metadata.get('dspy_context', ''))[:200],
            })
        
        # Upsert
        index.upsert(vectors=[(vector_id, embedding, metadata)])
        return True
    
    except Exception as e:
        print(f"⚠️  Failed to save: {e}")
        return False


def get_ingestion_state() -> Dict[str, Any]:
    """Load ingestion state"""
    if INGESTION_STATE_FILE.exists():
        with open(INGESTION_STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    return {
        'ingested_sheets': [],
        'ingested_entries': 0,
        'failed_entries': 0,
        'start_time': datetime.now().isoformat(),
        'last_update': None
    }


def save_ingestion_state(state: Dict[str, Any]) -> None:
    """Save ingestion state"""
    state['last_update'] = datetime.now().isoformat()
    
    DATA_DIR.mkdir(exist_ok=True)
    with open(INGESTION_STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def ingest_excel_data(genai_client: Any, pc: Pinecone, dry_run: bool = False,
                     resume: bool = False) -> None:
    """Main ingestion workflow"""
    
    print("\n" + "="*70)
    print("📊 Excel Data → Pinecone Ingestion")
    print("="*70 + "\n")
    
    wb = load_excel_workbook()
    text_data = load_raw_text_data()
    print(f"✅ Found {len(text_data)} text files\n")
    
    state = get_ingestion_state()
    ingested_count = 0
    failed_count = 0
    
    for sheet_idx, sheet_name in enumerate(wb.sheetnames[1:], start=1):
        
        if sheet_name in state['ingested_sheets'] and not resume:
            continue
        
        match = re.search(r'[0-9０-９]+\.\s*(.+)', sheet_name)
        character_name = match.group(1) if match else sheet_name
        
        print(f"[{sheet_idx:3d}/{len(wb.sheetnames)-1}] {character_name}")
        
        try:
            sheet = wb[sheet_name]
            sections = extract_excel_sections(sheet, character_name)
            
            # Generate metadata for each section
            section_metadata = {}
            for section_name, entries in sections.items():
                if entries:
                    preview = "\n".join(entries[:5])  # First 5 entries as preview
                    if not dry_run:
                        meta = generate_section_metadata(genai_client, character_name, section_name, preview)
                        section_metadata[section_name] = meta
                        print(f"    ✅ {section_name}")
            
            # Process each entry
            entry_idx = 0
            for section_name, entries in sections.items():
                for entry in entries:
                    if not entry or len(entry) < 5:
                        continue
                    
                    try:
                        text = format_technique_text(character_name, section_name, entry)
                        
                        if dry_run:
                            ingested_count += 1
                        else:
                            embedding = embed_text(genai_client, text)
                            if not embedding:
                                failed_count += 1
                                continue
                            
                            llm_meta = section_metadata.get(section_name, {})
                            success = save_to_pinecone(
                                pc, character_name, section_name,
                                entry, embedding, entry_idx, llm_metadata=llm_meta
                            )
                            
                            if success:
                                ingested_count += 1
                                # 進捗表示
                                if ingested_count % 5 == 0:
                                    print(f"      📊 進捗: {ingested_count}エントリ完了")
                                
                                # 10エントリごとに状態保存
                                if ingested_count % 10 == 0:
                                    state['ingested_entries'] = ingested_count
                                    state['failed_entries'] = failed_count
                                    save_ingestion_state(state)
                                    print(f"      💾 状態保存: {ingested_count}エントリ")
                            else:
                                failed_count += 1
                    
                    except Exception as e:
                        failed_count += 1
                    
                    entry_idx += 1
            
            state['ingested_sheets'].append(sheet_name)
            state['ingested_entries'] = ingested_count
            state['failed_entries'] = failed_count
            print(f"    ✅ 完了: {len([e for s in sections.values() for e in s])}エントリ処理")
        
        except Exception as e:
            print(f"  ❌ {e}")
            failed_count += 1
    
    if not dry_run:
        save_ingestion_state(state)
    
    print("\n" + "="*70)
    print(f"✅ Ingested: {ingested_count} | ❌ Failed: {failed_count}")
    print(f"📋 Sheets: {len(state['ingested_sheets'])}")
    print("="*70 + "\n")


def main():
    """Main entry point"""
    
    dry_run = '--dry-run' in sys.argv
    resume = '--resume' in sys.argv
    
    try:
        # Check if file exists
        if not EXCEL_FILE.exists():
            print(f"❌ Excel file not found: {EXCEL_FILE}")
            sys.exit(1)
        
        # Initialize APIs (skip in dry-run)
        if not dry_run:
            print("🔌 Initializing APIs...")
            genai_client, pc = initialize_apis()
            print("✅ APIs initialized\n")
        else:
            genai_client, pc = None, None
        
        # Run ingestion
        ingest_excel_data(genai_client, pc, dry_run=dry_run, resume=resume)
        
        print("🎉 Done!")
    
    except KeyboardInterrupt:
        print("\n⏸️  Interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"❌ Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
