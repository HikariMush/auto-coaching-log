#!/usr/bin/env python3
"""
マリオのExcelデータ抽出テスト
Pineconeに格納される形式を画面に表示
"""

import sys
from pathlib import Path
import openpyxl
import json

# Configuration
EXCEL_FILE = Path('src/brain/raw_data/スマブラSP フレームデータ by検証窓.xlsx')


def extract_excel_sections(sheet, character_name):
    """シートからセクション別にデータを抽出"""
    
    sections = {
        '行動フレーム': [],
        '能力値': [],
        '滞空フレーム': [],
    }
    
    section_keywords = {
        '行動フレーム': '行動フレーム',
        '能力値': '能力値',
        '滞空フレーム': '滞空フレーム',
    }
    
    current_section = None
    current_buffer = []
    
    # Scan all rows
    for row_idx in range(1, min(sheet.max_row + 1, 200)):  # 最初の200行
        row_data = []
        
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row_idx, col_idx)
            if cell.value is None:
                continue
            
            value = str(cell.value)
            
            # Skip formulas and images
            if value.startswith("='") or value.startswith('=') or '<' in value:
                continue
            
            row_data.append(value.strip())
        
        if not row_data:
            continue
        
        row_text = ' | '.join(row_data)
        
        # Check for section header
        for section_key, header in section_keywords.items():
            if header in row_text:
                # Save previous buffer
                if current_buffer and current_section:
                    sections[current_section].append(' | '.join(current_buffer))
                
                current_section = section_key
                current_buffer = []
                break
        else:
            # Add to current buffer
            if current_section and len(row_data) >= 2:
                current_buffer.extend(row_data)
                
                # If buffer is getting long, flush it
                if len(current_buffer) > 20:
                    sections[current_section].append(' | '.join(current_buffer))
                    current_buffer = []
    
    # Flush remaining buffer
    if current_buffer and current_section:
        sections[current_section].append(' | '.join(current_buffer))
    
    # Clean up sections
    for section in sections:
        sections[section] = [s.strip() for s in sections[section] if s.strip()]
    
    return sections


def format_technique_text(character, section, data):
    """技データを整形されたテキストに変換"""
    lines = [
        f"【キャラクター】{character}",
        f"【カテゴリ】{section}",
        f"【データ】{data[:500]}",  # 最初の500文字
    ]
    return "\n".join(lines)


def create_pinecone_vector_format(character, section, data, entry_idx):
    """Pinecone格納形式を作成（ベクトルIDとメタデータ）"""
    
    # ベクトルID生成
    char_clean = character.replace(' ', '_').replace('・', '-')
    section_clean = section.replace(' ', '_')
    vector_id = f"excel_{char_clean}_{section_clean}_{entry_idx}"
    
    # テキスト整形
    text = format_technique_text(character, section, data)
    
    # メタデータ
    metadata = {
        'character': character,
        'section': section,
        'source': 'excel_ingestion',
        'data_preview': data[:300],
        'entry_index': entry_idx,
    }
    
    return {
        'vector_id': vector_id,
        'text': text,
        'metadata': metadata,
        'embedding': '[768次元ベクトル（実際はGemini APIで生成）]'
    }


def main():
    """メイン処理"""
    
    print("\n" + "="*80)
    print("🎮 マリオのExcelデータ抽出テスト")
    print("="*80 + "\n")
    
    # Excelファイル読み込み
    if not EXCEL_FILE.exists():
        print(f"❌ Excel file not found: {EXCEL_FILE}")
        sys.exit(1)
    
    print(f"📂 Loading: {EXCEL_FILE.name}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    print(f"✅ Found {len(wb.sheetnames)} sheets\n")
    
    # マリオのシート探索
    mario_sheet = None
    for sheet_name in wb.sheetnames:
        if 'マリオ' in sheet_name and 'ドクター' not in sheet_name:
            mario_sheet = sheet_name
            break
    
    if not mario_sheet:
        print("❌ マリオのシートが見つかりません")
        sys.exit(1)
    
    print(f"📄 Processing sheet: {mario_sheet}\n")
    
    # データ抽出
    sheet = wb[mario_sheet]
    sections = extract_excel_sections(sheet, "マリオ")
    
    # 結果表示
    print("="*80)
    print("📊 抽出結果サマリー")
    print("="*80 + "\n")
    
    total_entries = 0
    for section_name, entries in sections.items():
        count = len(entries)
        total_entries += count
        print(f"  📍 {section_name}: {count} エントリ")
    
    print(f"\n  合計: {total_entries} エントリ\n")
    
    # Pinecone格納形式の例を表示
    print("="*80)
    print("🔍 Pinecone格納形式（全エントリ）")
    print("="*80 + "\n")
    
    entry_count = 0
    displayed_count = 0
    
    for section_name, entries in sections.items():
        if not entries:
            continue
        
        print(f"\n{'─'*80}")
        print(f"📦 セクション: {section_name}")
        print(f"{'─'*80}\n")
        
        for entry_idx, entry in enumerate(entries):
            if not entry or len(entry) < 5:
                continue
            
            # 全エントリを表示
            
            vector_format = create_pinecone_vector_format(
                "マリオ", section_name, entry, entry_idx
            )
            
            print(f"【エントリ {displayed_count + 1}】")
            print(f"Vector ID: {vector_format['vector_id']}")
            print(f"\nテキスト内容:")
            print("─" * 60)
            print(vector_format['text'])
            print("─" * 60)
            print(f"\nメタデータ:")
            print(json.dumps(vector_format['metadata'], ensure_ascii=False, indent=2))
            print(f"\nEmbedding: {vector_format['embedding']}")
            print("\n" + "="*80 + "\n")
            
            displayed_count += 1
            entry_count += 1
    
    # 統計情報
    print("="*80)
    print("📈 統計情報")
    print("="*80 + "\n")
    
    print(f"  キャラクター: マリオ")
    print(f"  処理可能エントリ数: {total_entries}")
    print(f"  表示したエントリ数: {displayed_count}")
    print(f"\n  推定Token数（1エントリあたり）: 300-400 tokens")
    print(f"  推定Token数（全エントリ）: {total_entries * 350:,} tokens")
    print(f"  推定埋め込みコスト: $0.00 (無料枠内)")
    print(f"  推定Pineconeコスト: ${total_entries * 0.0001:.4f}/月")
    
    print("\n" + "="*80)
    print("✅ テスト完了")
    print("="*80 + "\n")
    
    # 実行コマンド例
    print("💡 実際の取り込みコマンド例:")
    print("─" * 80)
    print("# マリオのみ取り込み")
    print("python -m src.utils.ingest_excel_data --start 1 --end 1 --embedding-delay 1.0")
    print("\n# 最初の3キャラ取り込み")
    print("python -m src.utils.ingest_excel_data --start 1 --end 3 --embedding-delay 1.0")
    print("─" * 80 + "\n")


if __name__ == '__main__':
    main()
